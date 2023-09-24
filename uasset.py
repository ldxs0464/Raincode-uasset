import openpyxl
import os
def bytereplace(bytes_data,new_bytes,pos):
    bytes_data=bytes_data[:pos]+new_bytes+bytes_data[pos+len(new_bytes):]
    return bytes_data

def printbytes(byte_data):
    for byte in byte_data:
        print(f'{byte:02X}',end=" ")

def tobytes(byte_data):
    data=""
    for byte in byte_data:
        data=data+format(byte,'02X')
    return data

def getfileName(input_string):
    if len(input_string)>0:
        last_slash_index = input_string.rfind('/')
        extracted_string = input_string[last_slash_index + 1:]
        return extracted_string
    else:
        return input_string
def repackUassetText(filepath):
    stringData=[]
    data_header = []
    data_pos_start = []
    data_pos_end = []
    dataheadersize = 19
    endSize = 4

    #import mods
    exportInTxt=True
    exportInExcel=False
    convertLineFeed=True

    with open('%s.txt' % os.path.splitext(filepath)[0], 'r', encoding='utf-16le') as file:
        for item in file:
            item= item.replace("\x0a", "")
            item = item.replace("\ufeff", "")
            item = item.replace("|", "\x0D\x0A")
            stringData.append(item)

    with (open(filepath, "rb") as file):
        bytes_data = file.read()
        #go to data position...
        data_position = int.from_bytes(bytes_data[52:56], byteorder='little')
        #data count(how many files)
        data_num=int.from_bytes(bytes_data[data_position+14:data_position+18], byteorder='little')
        pos=data_position

        if len(stringData) != data_num:
            print("data length is not match,so append a blank one...")
            for i in range(0,data_num-len(stringData)):
                stringData.append("")
                print("append a blank one...")

        data_header = bytes_data[pos:pos+18]
        #sheet['A1']=tobytes(bytes_data[pos:pos+18])

        # for data header skip...
        pos=pos+18
        for i in range(0,data_num):
            print(i,"/",data_num-1,end=" : ")

            header=bytes_data[pos:pos + dataheadersize]
            printbytes(header)
            data_pos_start.append(bytes_data[pos:pos + dataheadersize])
            pos=pos+dataheadersize

            #sheetpos = "A" + str(i + 2)
            #sheet[sheetpos] = getfileName(filepath)

            #sheetpos = "B" + str(i + 2)
            #sheet[sheetpos] = str(i)

            #sheetpos="C"+str(i+2)
            #sheet[sheetpos]= tobytes(header)
            #if saveInAction:
            #    workbook.save('%s.xlsx' % os.path.splitext(filepath)[0])

            data_len = int.from_bytes(bytes_data[pos:pos + 4], byteorder='little', signed=True)
            isUnicode=False

            if data_len<10000:
                #unicode uses negative values...
                if data_len<0:
                    data_len=-data_len;
                    #print("Unicode String!")
                    data_len=data_len*2
                    isUnicode=True
                #data length header skip
                pos=pos+4


                #add length for next line
                pos = pos + data_len
            else:
                pass



            #skip 38 5B 01 00
            #calculate length...
            endbyte = int.from_bytes(bytes_data[pos:pos + 4], byteorder='little', signed=False)
            if endbyte < 100:
                endbyte = endbyte + 4 +endSize
            else:
                endbyte=endSize

            # position editing...

            unk1 = int(header[10])
            match unk1:
                case 0x1c:
                    endbyte=0
                case 0x18:
                    endbyte=endbyte-4
                case 0x1A:
                    endbyte=0
                    pass
                case 0x10:

                    pass

            pass

            #sheetpos = "E" + str(i + 2)
            #sheet[sheetpos]= tobytes(bytes_data[pos:pos + endbyte])
            data_pos_end.append(bytes_data[pos:pos + endbyte])
            print(" : ", end="")
            printbytes(bytes_data[pos:pos + endbyte])


            pos=pos+endbyte

            print()
        #print(bytes_data)
    string_in_bytes = len(data_header)
    for i in range(0,data_num):
        string_in_bytes+=len(data_pos_start[i])
        data_for_string = stringData[i].encode("utf-16")
        data_for_string = data_for_string[2:]
        string_in_bytes +=len(data_for_string)
        if len(stringData[i]) != 0:
            string_in_bytes+=6
        string_in_bytes +=len(data_pos_end[i])
    data_size_head = int.from_bytes(bytes_data[44:48], byteorder='little')
    string_in_bytes=string_in_bytes-4
    bytes_data = bytereplace(bytes_data, string_in_bytes.to_bytes(4, byteorder='little'), (data_size_head + 8))

    with open('%s.uasset' % os.path.splitext(filepath)[0], 'wb') as file:

        file.write(bytes_data[0:data_position])
        file.write(data_header[0:len(data_header)])
        for i in range(0, data_num):
            file.write(data_pos_start[i])
            data_for_string = stringData[i].encode("utf-16")
            data_for_string = data_for_string[2:]

            if len(stringData[i]) > 0:
                file.write((-(len(stringData[i]) + 1)).to_bytes(4, byteorder='little', signed=True))
            file.write(data_for_string)
            if len(stringData[i]) != 0:
                file.write(b'\x00')
                file.write(b'\x00')
            file.write(data_pos_end[i])

    pass
def convertUassetText(filepath):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    stringData=[]
    dataheadersize = 19
    endSize = 4

    #export mods
    saveInAction=False
    exportInTxt=True
    exportInExcel=False
    convertLineFeed=True
    with (open(filepath, "rb") as file):
        bytes_data = file.read()
        #go to data position...
        data_position = int.from_bytes(bytes_data[52:56], byteorder='little')
        #data count(how many files)
        data_num=int.from_bytes(bytes_data[data_position+14:data_position+18], byteorder='little')
        pos=data_position
        data = ""
        sheet['A1']=tobytes(bytes_data[pos:pos+18])

        # for data header skip...
        pos=pos+18
        for i in range(0,data_num):
            data = ""
            print(i,"/",data_num-1,end=" : ")

            header=bytes_data[pos:pos + dataheadersize]
            printbytes(header)

            pos=pos+dataheadersize

            sheetpos = "A" + str(i + 2)
            sheet[sheetpos] = getfileName(filepath)

            sheetpos = "B" + str(i + 2)
            sheet[sheetpos] = str(i)

            sheetpos="C"+str(i+2)
            sheet[sheetpos]= tobytes(header)
            if saveInAction:
                workbook.save('%s.xlsx' % os.path.splitext(filepath)[0])

            data_len = int.from_bytes(bytes_data[pos:pos + 4], byteorder='little', signed=True)
            isUnicode=False

            if data_len<10000:
                #unicode uses negative values...
                if data_len<0:
                    data_len=-data_len;
                    #print("Unicode String!")
                    data_len=data_len*2
                    isUnicode=True
                #data length header skip
                pos=pos+4

                if isUnicode:
                    data = bytes_data[pos:pos + data_len - 2]
                    data = data.decode('utf-16le')
                else:
                    data = bytes_data[pos:pos + data_len - 1]
                    data = data.decode("ascii")

                if convertLineFeed:
                    data = data.replace('\x0d\x0a', '|')
                if len(data) == 0:
                    stringData.append(None)
                stringData.append(data)
                sheetpos = "D" + str(i + 2)
                sheet[sheetpos] = data

                if saveInAction:
                    workbook.save('%s.xlsx' % os.path.splitext(filepath)[0])

                #add length for next line
                pos = pos + data_len
            else:
                stringData.append(None)
                pass



            #skip 38 5B 01 00
            #calculate length...
            endbyte = int.from_bytes(bytes_data[pos:pos + 4], byteorder='little', signed=False)
            if endbyte < 100:
                endbyte = endbyte + 4 +endSize
            else:
                endbyte=endSize

            # position editing...

            unk1 = int(header[10])
            match unk1:
                case 0x1c:
                    endbyte=0
                case 0x18:
                    endbyte=endbyte-4
                case 0x1A:
                    endbyte=0
                    pass

            pass

            sheetpos = "E" + str(i + 2)
            sheet[sheetpos]= tobytes(bytes_data[pos:pos + endbyte])
            if saveInAction:
                workbook.save('%s.xlsx' % os.path.splitext(filepath)[0])


            pos=pos+endbyte
            pass
            print(data)
        #print(bytes_data)
    if pos != len(bytes_data):
        print("Length is not matched(Check the code!)")
        input()
    if not saveInAction:
        if exportInExcel:
            workbook.save('%s.xlsx' % os.path.splitext(filepath)[0])
    if exportInTxt:
        with open('%s.txt'%os.path.splitext(filepath)[0], 'w',encoding='utf-16le') as file:
            file.write('\ufeff')
            for i in range(0,len(stringData)):
                if i == len(stringData)-1:
                    if stringData[i] is None:
                        file.write('')
                    else:
                        file.write(stringData[i])
                else:
                    if stringData[i] is None:
                        file.write('\n')
                    else:
                        file.write(stringData[i]+'\n')
#convertUassetText("RCSFL.uasset")
def extractinsizefolder(folder_path):
    fileCount=0
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.uasset'):
                file_path = os.path.join(root, file)
                print(file)
                convertUassetText(file_path)
                fileCount+=1
    print("files=",fileCount)

def importinsizefolder(folder_path):
    fileCount=0
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.uasset'):
                file_path = os.path.join(root, file)
                print(file)
                repackUassetText(file_path)
                fileCount+=1
    print("files=",fileCount)

def remove_file_extension(file_name):
    base_name = os.path.basename(file_name)
    name_without_extension = os.path.splitext(base_name)[0]
    return name_without_extension
def convert2XlsxAll(folder_path):
    # 폴더 내의 파일 목록을 가져옵니다.
    file_data = {}

    for root, dirs, files in os.walk(folder_path):
        files=sorted(files)
        for file in files:
            if file.endswith(".txt"):
                file_path = os.path.join(root, file)
                folder_name = os.path.basename(os.path.dirname(file_path))
                #만약 폴더 내부의 폴더면 sheet name을 a/b 식으로 진행한다
                print(file," : ",file_path.count("/"))
                if file_path.count("/")>2:
                    folder_name = root.split('/')[1:]
                    folder_name = '/'.join(folder_name)
                    folder_name = folder_name.replace("/","_")
                    print(folder_name)

                if folder_name not in file_data:
                    file_data[folder_name] = []
                file_data[folder_name].append(file_path)

    #print(file_data)
    # 엑셀 파일 생성
    wb = openpyxl.Workbook()

    # 각 폴더마다 시트 생성하고 데이터 입력
    for folder_name, files in file_data.items():
        ws = wb.create_sheet(title=folder_name)
        ws.append(["파일명", "줄 번호", "내용"])

        for file_path in files:
            with open(file_path, 'r', encoding='utf-16le') as file:
                lines = file.readlines()
                for line_num, line in enumerate(lines, start=1):
                    line=line.replace("\n", "")
                    ws.append([remove_file_extension(os.path.basename(file_path)), line_num, line])

    # 기본 시트 제거
    default_sheet = wb.get_sheet_by_name("Sheet")
    wb.remove(default_sheet)

    # 엑셀 파일 저장
    output_file =  folder_path+".xlsx"
    wb.save(output_file)
    print(f"엑셀 파일이 {output_file}에 저장되었습니다.")

def deleteAllTxt(file_path):
    for root, dirs, files in os.walk(file_path):
        for file in files:
            if file.endswith(".txt"):
                file_to_delete = os.path.join(root, file)
                os.remove(file_to_delete)
def xlsx2Txt(file_path):
    deleteAllTxt(remove_file_extension(file_path))
    # 엑셀 파일 열기
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # 각 시트별로 작업
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        sheetName=sheet
        sheetName=sheetName.replace("_","/")

        output_folder = remove_file_extension(file_path)+"/"+sheetName#os.path.join(os.path.dirname(file_path), sheet)

        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        for row in ws.iter_rows(min_row=2, values_only=True):
            file_name = row[0]  # A열에 있는 파일명
            file_num= row[1]
            content = row[2]  # C열에 있는 내용

            # 파일명으로 txt 파일 경로 구성
            txt_file_path = os.path.join(output_folder, f"{file_name}.txt")

            if not os.path.isfile(txt_file_path):
                with open(txt_file_path, "wb") as file:
                    # BOM (UTF-16 LE) 추가
                    file.write(b'\xff\xfe')

            # 텍스트 파일로 데이터 쓰기 (UTF-16LE)
            with open(txt_file_path, 'a', encoding='utf-16le') as txt_file:
                if file_num != 1:
                    txt_file.write('\n')
                if content == None:
                    pass
                else:
                    txt_file.write(str(content))



#xlsx2Txt("test.xlsx")


# 테스트
#folder_path = "Texts"  # 입력하려는 폴더 경로
#convert2XlsxAll(folder_path)



# Example usage
#txt2Excel("TextsJP/design")
#xlsxinsizefolder("TextsJaJP")
#extractinsizefolder("TextsJaJP")
#convertUassetText("RCTrophy.uasset")
#importinsizefolder("Texts")
#repackUassetText("RCInvestigationName.uasset")
folder_name = "Texts"
while True:

    print("1.uasset->txt")
    print("2.txt->xlsx")
    print("3.xlsx->txt")
    print("4.txt->uasset")
    print("5.clear all txt")
    print("6.[DEBUG]uasset->txt")
    print("7.[DEBUG]txt->uasset")
    print("8.[DEBUG]change folder name (current :",folder_name,")")
    print("Choose your option:",end="")

    choice = int(input())
    if choice == 1:
        print("Are you sure? type yes for process")
        answer = input()
        if answer == "yes":
            extractinsizefolder(folder_name)
    elif choice == 2:
        print("Are you sure? type yes for process")
        answer = input()
        if answer == "yes":
            convert2XlsxAll(folder_name)
    elif choice == 3:
        xlsx2Txt(folder_name + ".xlsx")
    elif choice == 4:
        importinsizefolder(folder_name)
    elif choice == 5:
        deleteAllTxt(folder_name)
    elif choice == 6:
        print("Please input full path of uasset file:", end="")
        path = str(input())
        convertUassetText(path)
    elif choice == 7:
        print("Please input full path of uasset file:", end="")
        path = str(input())
        repackUassetText(path)
    elif choice == 8:
        print("Please input new folder name:", end="")
        folder_name=str(input())